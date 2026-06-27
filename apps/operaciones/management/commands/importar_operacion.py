"""Importa operación real desde las planillas Retiro.xlsx y Entregas.xlsx.

El N° de CONTENEDOR es la llave de negocio que enlaza ambas planillas:
  - Retiro    → sacar el contenedor del puerto (fecha_retiro, estado_retiro).
  - Entregas  → llevarlo al cliente y devolver el vacío (fecha_entrega, etc.).

El comando:
  1. Lee ambas planillas (openpyxl), tolerando cabeceras con espacios/acentos.
  2. Filtra a una ventana reciente (--dias) y limita el total (--max).
  3. Fusiona retiro + entrega por contenedor en una ETA enriquecida.
  4. Crea catálogos (cliente, conductor, camión, contenedor, puerto) y
     movimientos de trazabilidad.

Uso típico:
    python manage.py importar_operacion --reset --dias 30 --max 400
"""

from __future__ import annotations

import os
import random
import unicodedata
from datetime import date, datetime, time, timedelta

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone

from apps.operaciones.models import (
    AgentePortuario,
    Camion,
    Cliente,
    Conductor,
    Contenedor,
    ETA,
    Movimiento,
)

BASE_DOCS = os.path.join(os.getcwd(), "docs")


# ---------------------------------------------------------------------------
# Utilidades de lectura / normalización
# ---------------------------------------------------------------------------
def _norm(texto) -> str:
    """Normaliza una cabecera/valor: sin acentos, mayúsculas, sin espacios extra."""
    if texto is None:
        return ""
    s = str(texto)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.upper().split()).strip()


def _clave_contenedor(valor) -> str:
    """Llave de contenedor: sin espacios internos, mayúsculas. Trunca a 15."""
    if valor is None:
        return ""
    s = str(valor).upper()
    s = "".join(s.split())
    return s[:15]


def _a_fecha(valor):
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    if isinstance(valor, str) and valor.strip():
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(valor.strip()[:10], fmt).date()
            except ValueError:
                continue
    return None


def _a_hora(valor):
    if isinstance(valor, datetime):
        return valor.time()
    if isinstance(valor, time):
        return valor
    if isinstance(valor, str) and valor.strip():
        for fmt in ("%H:%M:%S", "%H:%M"):
            try:
                return datetime.strptime(valor.strip()[:8], fmt).time()
            except ValueError:
                continue
    return None


def _a_entero(valor):
    if valor in (None, ""):
        return None
    try:
        return int(float(str(valor).replace(".", "").replace(",", ".")))
    except (ValueError, TypeError):
        try:
            return int(float(valor))
        except (ValueError, TypeError):
            return None


def _leer_hoja(ruta, hoja):
    """Devuelve (lista_de_dicts_normalizados). Claves = cabeceras normalizadas."""
    from openpyxl import load_workbook

    wb = load_workbook(ruta, read_only=True, data_only=True)
    if hoja not in wb.sheetnames:
        # primera hoja por defecto
        ws = wb[wb.sheetnames[0]]
    else:
        ws = wb[hoja]

    filas = ws.iter_rows(values_only=True)
    try:
        cabeceras = next(filas)
    except StopIteration:
        return []
    columnas = [_norm(c) for c in cabeceras]

    registros = []
    for fila in filas:
        if fila is None:
            continue
        d = {}
        for col, val in zip(columnas, fila):
            if col:
                d[col] = val
        if any(v not in (None, "") for v in d.values()):
            registros.append(d)
    wb.close()
    return registros


def _get(d, *claves):
    """Primer valor no vacío entre varias cabeceras candidatas (normalizadas)."""
    for k in claves:
        if k in d and d[k] not in (None, ""):
            return d[k]
    return None


# ---------------------------------------------------------------------------
# Mapeos a choices del modelo
# ---------------------------------------------------------------------------
def _map_estado_retiro(valor):
    n = _norm(valor)
    if "RETIR" in n:
        return ETA.EstadoRetiro.RETIRADO
    if "PEND" in n:
        return ETA.EstadoRetiro.PENDIENTE
    if "CANCEL" in n:
        return ETA.EstadoRetiro.CANCELADO
    if "REPROG" in n:
        return ETA.EstadoRetiro.REPROGRAMADO
    return ""


def _map_estado_entrega(valor):
    n = _norm(valor)
    if "NO ENTREG" in n:
        return ETA.EstadoEntrega.NO_ENTREGADO
    if "ENTREG" in n:
        return ETA.EstadoEntrega.ENTREGADO
    if "SIN LLEG" in n:
        return ETA.EstadoEntrega.SIN_LLEGADA
    if "REPROG" in n:
        return ETA.EstadoEntrega.REPROGRAMADO
    if "PEND" in n:
        return ETA.EstadoEntrega.PENDIENTE
    return ""


def _map_otd(valor):
    n = _norm(valor)
    if "ON TIME" in n:
        return ETA.OTD.ON_TIME
    if "OFF TIME" in n:
        return ETA.OTD.OFF_TIME
    if "SIN" in n:
        return ETA.OTD.SIN_LLEGADA
    return ""


def _map_operacion(valor):
    n = _norm(valor)
    if "IMPO" in n:
        return ETA.Operacion.IMPORTACION
    if "EXPO" in n:
        return ETA.Operacion.EXPORTACION
    return ""


def _map_tipo_contenedor(dimension):
    n = _norm(dimension)
    if "REEF" in n or "RF" in n:
        return Contenedor.Tipo.REEFER
    if "40" in n and "HC" in n:
        return Contenedor.Tipo.HC_40
    if "40" in n:
        return Contenedor.Tipo.DRY_40
    if "20" in n:
        return Contenedor.Tipo.DRY_20
    return Contenedor.Tipo.OTRO


class Command(BaseCommand):
    help = "Importa operación real desde docs/Retiro.xlsx y docs/Entregas.xlsx."

    def add_arguments(self, parser):
        parser.add_argument("--dias", type=int, default=30,
                            help="Ventana de días hacia atrás desde la fecha más reciente.")
        parser.add_argument("--max", type=int, default=400,
                            help="Máximo de ETAs a crear.")
        parser.add_argument("--reset", action="store_true",
                            help="Elimina ETAs/movimientos/catálogos operativos antes de cargar.")
        parser.add_argument("--retiro", default=os.path.join(BASE_DOCS, "Retiro.xlsx"))
        parser.add_argument("--entregas", default=os.path.join(BASE_DOCS, "Entregas.xlsx"))

    # ------------------------------------------------------------------
    def handle(self, *args, **options):
        ruta_ret = options["retiro"]
        ruta_ent = options["entregas"]
        for ruta in (ruta_ret, ruta_ent):
            if not os.path.exists(ruta):
                raise CommandError(f"No se encontró la planilla: {ruta}")

        self.stdout.write("Leyendo planillas...")
        retiros = _leer_hoja(ruta_ret, "RETIROS")
        entregas = _leer_hoja(ruta_ent, "ENTREGAS")
        self.stdout.write(f"  Retiro.xlsx:   {len(retiros)} filas")
        self.stdout.write(f"  Entregas.xlsx: {len(entregas)} filas")

        registros = self._construir_registros(retiros, entregas, options["dias"], options["max"])
        self.stdout.write(f"Registros fusionados en ventana: {len(registros)}")

        if options["reset"]:
            self._reset()

        creadas = self._cargar(registros)
        self.stdout.write(self.style.SUCCESS(
            f"Importación completa: {creadas} ETAs cargadas."
        ))

    # ------------------------------------------------------------------
    def _construir_registros(self, retiros, entregas, dias, maximo):
        """Fusiona retiro + entrega por contenedor y filtra a la ventana."""
        # Parseo de retiros
        ret_parsed = []
        for r in retiros:
            cont = _clave_contenedor(_get(r, "CONTENEDOR"))
            if not cont:
                continue
            f_ret = _a_fecha(_get(r, "F RETIRO", "FECHA RETIRO"))
            ret_parsed.append({"cont": cont, "f_ret": f_ret, "raw": r})

        ent_parsed = []
        for e in entregas:
            cont = _clave_contenedor(_get(e, "CONTENEDOR"))
            if not cont:
                continue
            f_ent = _a_fecha(_get(e, "FECHA ENTREGA"))
            ent_parsed.append({"cont": cont, "f_ent": f_ent, "raw": e})

        # Fecha máxima de referencia
        fechas = [x["f_ret"] for x in ret_parsed if x["f_ret"]]
        fechas += [x["f_ent"] for x in ent_parsed if x["f_ent"]]
        if not fechas:
            raise CommandError("No se pudieron parsear fechas en las planillas.")
        fmax = max(fechas)
        corte = fmax - timedelta(days=dias)

        # Filtrar a ventana, ordenar por fecha desc
        ret_win = sorted(
            [x for x in ret_parsed if x["f_ret"] and x["f_ret"] >= corte],
            key=lambda x: x["f_ret"], reverse=True,
        )
        ent_win = sorted(
            [x for x in ent_parsed if x["f_ent"] and x["f_ent"] >= corte],
            key=lambda x: x["f_ent"], reverse=True,
        )

        # Índice de entregas por contenedor (la más reciente gana)
        ent_por_cont = {}
        for e in ent_win:
            ent_por_cont.setdefault(e["cont"], e)

        registros = []
        usados_ent = set()

        # 1) Retiros (enriquecidos con su entrega si existe)
        for r in ret_win:
            if len(registros) >= maximo:
                break
            ent = ent_por_cont.get(r["cont"])
            if ent:
                usados_ent.add(r["cont"])
            registros.append(self._fusionar(r["cont"], r, ent))

        # 2) Entregas sin retiro en la ventana
        for cont, e in ent_por_cont.items():
            if len(registros) >= maximo:
                break
            if cont in usados_ent:
                continue
            registros.append(self._fusionar(cont, None, e))

        return registros

    # ------------------------------------------------------------------
    def _fusionar(self, cont, r, e):
        """Construye un dict de ETA fusionando retiro (r) y entrega (e)."""
        rr = r["raw"] if r else {}
        ee = e["raw"] if e else {}

        cliente = _get(rr, "CLIENTE") or _get(ee, "CLIENTE") or "Sin cliente"
        despacho = _get(rr, "DESPACHO") or _get(ee, "DESPACHO") or ""
        dimension = _get(rr, "DIMENSION") or _get(ee, "TIPO") or ""
        puerto = _get(rr, "PUERTO") or _get(ee, "LUGAR RETIRO") or ""
        nave = _get(rr, "M/N") or ""
        peso = _a_entero(_get(rr, "PESO") or _get(ee, "PESO"))
        operacion = _map_operacion(_get(ee, "OPERACION"))

        chofer = _get(rr, "CHOFER RETIRO") or _get(ee, "CONDUCTOR") or ""
        camion = _get(rr, "CAMION") or _get(ee, "EQUIPO") or ""

        f_ret = r["f_ret"] if r else None
        f_ent = e["f_ent"] if e else None
        horario = _a_hora(_get(rr, "HORARIO")) or _a_hora(_get(ee, "HORA ENTREGA"))

        estado_retiro = _map_estado_retiro(_get(rr, "ESTADO DE RETIRO"))
        estado_entrega = _map_estado_entrega(_get(ee, "ESTADO DE ENTREGA"))
        otd = _map_otd(_get(ee, "OTD") or _get(rr, "OTD"))

        direccion = _get(ee, "DIRECCION ENTREGA") or ""
        deposito_dev = _get(ee, "DEPOSITO DEVOLUCION") or ""
        observ = _get(rr, "OBSERVACIONES") or _get(ee, "NOTA") or ""

        return {
            "contenedor": cont,
            "cliente": str(cliente).strip()[:150],
            "despacho": str(despacho).strip()[:30],
            "dimension": str(dimension).strip()[:30],
            "puerto": str(puerto).strip()[:40],
            "nave": str(nave).strip()[:80],
            "peso": peso,
            "operacion": operacion,
            "chofer": str(chofer).strip()[:150] if chofer else "",
            "camion": str(camion).strip()[:10] if camion else "",
            "f_ret": f_ret,
            "f_ent": f_ent,
            "horario": horario,
            "estado_retiro": estado_retiro,
            "estado_entrega": estado_entrega,
            "otd": otd,
            "direccion": str(direccion).strip()[:200] if direccion else "",
            "deposito_dev": str(deposito_dev).strip()[:80] if deposito_dev else "",
            "observ": str(observ).strip() if observ else "",
        }

    # ------------------------------------------------------------------
    def _reset(self):
        self.stdout.write("Eliminando datos operativos previos...")
        Movimiento.objects.all().delete()
        ETA.objects.all().delete()
        Contenedor.objects.all().delete()
        Conductor.objects.all().delete()
        Camion.objects.all().delete()
        Cliente.objects.all().delete()
        AgentePortuario.objects.all().delete()

    # ------------------------------------------------------------------
    def _estado_ciclo(self, reg):
        """Deriva el EstadoCiclo físico a partir de los estados documentales."""
        if reg["estado_entrega"] == ETA.EstadoEntrega.ENTREGADO:
            if reg["deposito_dev"]:
                return ETA.EstadoCiclo.DESPACHADO_PUERTO
            return ETA.EstadoCiclo.DESPACHADO_CLIENTE
        if reg["estado_retiro"] == ETA.EstadoRetiro.RETIRADO:
            return ETA.EstadoCiclo.ALMACENADO
        if reg["estado_retiro"] == ETA.EstadoRetiro.PENDIENTE:
            return ETA.EstadoCiclo.SOLICITADO
        return ETA.EstadoCiclo.ASIGNADO

    # ------------------------------------------------------------------
    @transaction.atomic
    def _cargar(self, registros):
        rnd = random.Random(2026)
        clientes_cache = {}
        conductores_cache = {}
        camiones_cache = {}
        puertos_cache = {}

        def cliente_de(nombre):
            if nombre not in clientes_cache:
                obj, _ = Cliente.objects.get_or_create(nombre=nombre)
                clientes_cache[nombre] = obj
            return clientes_cache[nombre]

        def puerto_de(nombre):
            nombre = nombre or "Sin puerto"
            if nombre not in puertos_cache:
                obj, _ = AgentePortuario.objects.get_or_create(
                    nombre=nombre, defaults={"sigla": nombre[:20]}
                )
                puertos_cache[nombre] = obj
            return puertos_cache[nombre]

        def conductor_de(nombre):
            if not nombre:
                return None
            if nombre not in conductores_cache:
                estado = (Conductor.Estado.INOPERATIVO
                          if rnd.random() < 0.17 else Conductor.Estado.ACTIVO)
                obj, creado = Conductor.objects.get_or_create(
                    nombre=nombre, defaults={"estado": estado}
                )
                conductores_cache[nombre] = obj
            return conductores_cache[nombre]

        def camion_de(patente):
            if not patente:
                return None
            if patente not in camiones_cache:
                obj, _ = Camion.objects.get_or_create(patente=patente)
                camiones_cache[patente] = obj
            return camiones_cache[patente]

        creadas = 0
        for i, reg in enumerate(registros, start=1):
            contenedor, _ = Contenedor.objects.get_or_create(
                codigo=reg["contenedor"],
                defaults={
                    "tipo": _map_tipo_contenedor(reg["dimension"]),
                    "estado": Contenedor.Estado.CARGADO,
                },
            )
            estado_ciclo = self._estado_ciclo(reg)
            fecha_base = reg["f_ret"] or reg["f_ent"] or timezone.now().date()

            eta = ETA.objects.create(
                numero=f"OPS-2026-{i:05d}",
                cliente=cliente_de(reg["cliente"]),
                agente=puerto_de(reg["puerto"]),
                contenedor=contenedor,
                conductor=conductor_de(reg["chofer"]),
                camion=camion_de(reg["camion"]),
                deposito=reg["deposito_dev"][:120],
                fecha=fecha_base,
                hora_retiro=reg["horario"],
                estado=estado_ciclo,
                observaciones=reg["observ"],
                # --- campos operativos reales ---
                despacho=reg["despacho"],
                operacion=reg["operacion"],
                nave=reg["nave"],
                puerto=reg["puerto"],
                dimension=reg["dimension"],
                peso=reg["peso"],
                direccion_entrega=reg["direccion"],
                deposito_devolucion=reg["deposito_dev"],
                fecha_retiro=reg["f_ret"],
                fecha_entrega=reg["f_ent"],
                horario=reg["horario"],
                estado_retiro=reg["estado_retiro"],
                estado_entrega=reg["estado_entrega"],
                otd=reg["otd"],
            )

            self._crear_movimientos(eta, reg)
            creadas += 1

        return creadas

    # ------------------------------------------------------------------
    def _crear_movimientos(self, eta, reg):
        hora = reg["horario"] or time(9, 0)
        if reg["f_ret"]:
            Movimiento.objects.create(
                eta=eta,
                tipo=Movimiento.Tipo.RETIRO,
                fecha=timezone.make_aware(datetime.combine(reg["f_ret"], hora)),
                observacion=f"Retiro en {reg['puerto']}".strip(),
            )
        if reg["estado_entrega"] == ETA.EstadoEntrega.ENTREGADO and reg["f_ent"]:
            Movimiento.objects.create(
                eta=eta,
                tipo=Movimiento.Tipo.DESPACHO_CLIENTE,
                fecha=timezone.make_aware(datetime.combine(reg["f_ent"], hora)),
                observacion=reg["direccion"][:200],
            )
            if reg["deposito_dev"]:
                Movimiento.objects.create(
                    eta=eta,
                    tipo=Movimiento.Tipo.DESPACHO_PUERTO,
                    fecha=timezone.make_aware(datetime.combine(reg["f_ent"], hora)),
                    observacion=f"Devolución vacío en {reg['deposito_dev']}",
                )
